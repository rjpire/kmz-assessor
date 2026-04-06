#!/usr/bin/env python3
"""
KMZ → County Assessor → Excel
=================================
Extracts geographic points (and samples along LineStrings) from a KMZ file,
queries free county assessor GIS services, and exports owner/parcel records
to Excel.

Handles both Point placemarks and LineString placemarks (e.g. pipeline routes).
For LineStrings, the line is sampled at a configurable interval and each parcel
is queried once — duplicates are detected by parcel_id and not re-queried.

Usage:
    # Pipeline KMZ with 100 m sampling (county GIS, no API key needed):
    python kmz_to_assessor.py pipeline.kmz --sample-interval 0.001

    # Point KMZ using every vertex:
    python kmz_to_assessor.py mypoints.kmz

Counties must be registered in counties.yaml (Weld County, CO included).
Use the Streamlit app's County Registry panel to add new counties.
"""

import argparse
import os
import sys
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    import requests
    from bs4 import BeautifulSoup
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependencies. Run:\n    pip install requests openpyxl beautifulsoup4")
    sys.exit(1)


# ---------------------------------------------------------------------------
# KMZ / KML Parsing
# ---------------------------------------------------------------------------

KML_NS = "http://www.opengis.net/kml/2.2"


def _find(element, tag, ns=KML_NS):
    """Find a child element with or without namespace."""
    result = element.find(f"{{{ns}}}{tag}")
    if result is None:
        result = element.find(tag)
    return result


def _findall(element, tag, ns=KML_NS):
    """Find all child elements with or without namespace."""
    results = element.findall(f".//{{{ns}}}{tag}")
    if not results:
        results = element.findall(f".//{tag}")
    return results


def _parse_description_html(html_text: str) -> dict:
    """
    Parse an HTML <table> description from a KML placemark into a flat dict.

    Expects rows like:
        <tr><td>Company</td><td>Acme Pipeline Co</td></tr>

    Returns {field_name: value} for every 2-cell row found.
    Returns {} on empty, non-HTML, or table-less input.
    """
    if not html_text or not html_text.strip():
        return {}
    try:
        soup = BeautifulSoup(html_text, "html.parser")
        result = {}
        for row in soup.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) == 2:
                key = cells[0].get_text(strip=True)
                val = cells[1].get_text(strip=True)
                if key:
                    result[key] = val
        return result
    except Exception:
        return {}


def _coords_from_linestring(coord_text: str) -> list:
    """
    Parse a KML <coordinates> text block into a list of (lat, lng) tuples.

    KML order is: longitude,latitude,altitude (altitude optional).
    Returns [] if coord_text is empty or unparseable.
    """
    result = []
    for token in (coord_text or "").strip().split():
        parts = token.split(",")
        if len(parts) < 2:
            continue
        try:
            lng = float(parts[0])
            lat = float(parts[1])
            result.append((lat, lng))
        except ValueError:
            continue
    return result


def _sample_linestring(coords: list, interval_deg: float) -> list:
    """
    Walk a polyline and emit sample points every `interval_deg` degrees of
    cumulative straight-line distance.

    interval_deg <= 0  → return all vertices unchanged.
    Always emits the first vertex. Uses Euclidean degree-space (adequate for
    spans of a single county; 0.001° ≈ 111 m north-south at any latitude).

    Returns list of (lat, lng) tuples.
    """
    if interval_deg <= 0 or len(coords) <= 1:
        return list(coords)

    sampled = [coords[0]]
    accumulated = 0.0

    for i in range(1, len(coords)):
        prev = coords[i - 1]
        curr = coords[i]
        dlat = curr[0] - prev[0]
        dlng = curr[1] - prev[1]
        seg_len = (dlat ** 2 + dlng ** 2) ** 0.5
        accumulated += seg_len

        if accumulated >= interval_deg:
            sampled.append(curr)
            accumulated = 0.0

    # Always include last vertex if it wasn't just added
    if sampled[-1] != coords[-1]:
        sampled.append(coords[-1])

    return sampled


def parse_kmz(filepath: str, sample_interval_deg: float = 0.0) -> list:
    """
    Extract placemarks from a KMZ or KML file.

    Handles both Point and LineString geometry (including LineStrings wrapped
    inside MultiGeometry).

    Point placemarks → one dict per placemark (original behaviour, unchanged).
    LineString placemarks → sampled into multiple dicts (one per sample point),
    each carrying the segment's metadata parsed from the description HTML.

    Parameters
    ----------
    filepath             : path to .kmz or .kml file
    sample_interval_deg  : spacing in decimal degrees for LineString sampling.
                           0.0 (default) = use every vertex.
                           0.001 ≈ 100 m (good for rural pipeline parcels).

    Returns list of dicts with keys:
        name, description, lat, lng   — present for all geometry types
        geometry_type                 — "Point" or "LineString"
        vertex_index                  — 0-based index within the placemark
        meta                          — dict from _parse_description_html
    """
    filepath = Path(filepath)

    if filepath.suffix.lower() == ".kmz":
        with zipfile.ZipFile(filepath) as kmz:
            kml_files = [f for f in kmz.namelist() if f.endswith(".kml")]
            if not kml_files:
                raise ValueError("No .kml file found inside the KMZ archive.")
            main_kml = "doc.kml" if "doc.kml" in kml_files else kml_files[0]
            with kmz.open(main_kml) as fh:
                tree = ET.parse(fh)
    elif filepath.suffix.lower() == ".kml":
        tree = ET.parse(filepath)
    else:
        raise ValueError(f"Unsupported file type: {filepath.suffix}  (expected .kmz or .kml)")

    root = tree.getroot()
    placemarks = _findall(root, "Placemark")

    points = []
    skipped = 0

    for pm in placemarks:
        # Common metadata for all geometry types
        name_el = _find(pm, "name")
        name = (name_el.text or "").strip() if name_el is not None else "Unnamed"

        desc_el = _find(pm, "description")
        raw_desc = (desc_el.text or "").strip() if desc_el is not None else ""
        meta = _parse_description_html(raw_desc)

        # ── Point ────────────────────────────────────────────────────────────
        point_el = _find(pm, "Point")
        if point_el is not None:
            coord_el = _find(point_el, "coordinates")
            if coord_el is None or not (coord_el.text or "").strip():
                skipped += 1
                continue

            parts = coord_el.text.strip().split(",")
            if len(parts) < 2:
                skipped += 1
                continue

            try:
                lng = float(parts[0])
                lat = float(parts[1])
            except ValueError:
                skipped += 1
                continue

            points.append({
                "name": name,
                "description": raw_desc,
                "lat": lat,
                "lng": lng,
                "geometry_type": "Point",
                "vertex_index": 0,
                "meta": meta,
            })
            continue

        # ── LineString (including those inside MultiGeometry) ─────────────
        # Use _findall so we pick up <LineString> nested under <MultiGeometry>
        line_els = _findall(pm, "LineString")
        if line_els:
            for line_el in line_els:
                coord_el = _find(line_el, "coordinates")
                if coord_el is None or not (coord_el.text or "").strip():
                    skipped += 1
                    continue
                raw_coords = _coords_from_linestring(coord_el.text)
                if not raw_coords:
                    skipped += 1
                    continue
                sampled = _sample_linestring(raw_coords, sample_interval_deg)
                for idx, (lat, lng) in enumerate(sampled):
                    points.append({
                        "name": name,
                        "description": raw_desc,
                        "lat": lat,
                        "lng": lng,
                        "geometry_type": "LineString",
                        "vertex_index": idx,
                        "meta": meta,
                    })
            continue

        # ── Unsupported (Polygon, NetworkLink, etc.) ──────────────────────
        skipped += 1

    if skipped:
        try:
            print(f"   Skipped {skipped} unsupported feature(s) (polygons, folders, etc.)")
        except UnicodeEncodeError:
            print(f"   Skipped {skipped} unsupported feature(s)")

    return points


# ---------------------------------------------------------------------------
# Deduplicated Lookup Loop — County GIS
# ---------------------------------------------------------------------------

def lookup_parcels_deduped(
    points: list,
    progress_callback=None,
    registry=None,
) -> list:
    """
    Run county GIS parcel lookups for all sampled points, skipping duplicates.

    Once a parcel_id has been seen, subsequent points that resolve to that
    same parcel copy the cached result (status = "DUPLICATE") without making
    another network call. This is critical for pipeline KMZ files where many
    sample points may fall on the same large rural parcel.

    Parameters
    ----------
    points            : output of parse_kmz()
    progress_callback : optional callable(i, total, label) called per point
    registry          : CountyRegistry instance (loads default counties.yaml if None)

    Returns list of row dicts ready for write_excel(), with extra keys:
        deduped  — True if this row reused a cached parcel result
    """
    from county_lookup import CountyRegistry, lookup_parcel_county_gis

    if registry is None:
        registry = CountyRegistry()

    seen_parcels: dict = {}   # parcel_id → parcel dict
    rows = []

    for i, pt in enumerate(points):
        label = pt.get("name") or f"Sample {i + 1}"
        if progress_callback:
            progress_callback(i, len(points), label)

        try:
            parcel = lookup_parcel_county_gis(pt["lat"], pt["lng"], registry)
            unregistered = parcel.pop("_unregistered_county", None)
            if unregistered:
                parcel = {}
                raw_status = f"No county registered: {unregistered}"
            else:
                raw_status = "Found" if parcel else "No record found"

            pid = parcel.get("parcel_id", "")

            if pid and pid in seen_parcels:
                parcel = seen_parcels[pid]
                status = "DUPLICATE"
                deduped = True
            elif pid:
                seen_parcels[pid] = parcel
                status = raw_status
                deduped = False
            else:
                status = raw_status
                deduped = False

        except Exception as e:
            parcel = {}
            status = f"ERROR: {e}"
            deduped = False

        rows.append({**pt, **parcel, "status": status, "deduped": deduped})

    return rows


# ---------------------------------------------------------------------------
# Meta Flattening
# ---------------------------------------------------------------------------

# Mapping from HTML description table field names → flat row keys used in COLUMNS
_META_KEY_MAP = {
    "LineName":          "line_name",
    "LineDesignator":    "line_designator",
    "BeginStation":      "begin_station",
    "EndStation":        "end_station",
    "Operator":          "operator",
    "ProductType":       "product_type",
    "OperationalStatus": "operational_status",
}


def _flatten_meta(row: dict) -> dict:
    """
    Promote known LineString metadata keys from row['meta'] into top-level
    row keys so write_excel() can find them by key name.

    Missing meta keys are set to "" (blank Excel cell).
    Works transparently for Point rows (meta={}).
    Returns the same dict (mutated in place) for convenience.
    """
    meta = row.get("meta") or {}
    for html_key, row_key in _META_KEY_MAP.items():
        row[row_key] = meta.get(html_key, "")
    return row


# ---------------------------------------------------------------------------
# Excel Output
# ---------------------------------------------------------------------------

COLUMNS = [
    # (header label, field key, col width)
    ("Point Name",          "name",               22),
    ("Latitude",            "lat",                13),
    ("Longitude",           "lng",                13),
    ("Geometry Type",       "geometry_type",      16),
    ("Parcel ID / APN",     "parcel_id",          24),
    ("Owner Name",          "owner",              30),
    ("Mailing Address",     "mail_address",       28),
    ("Mail City",           "mail_city",          18),
    ("Mail State",          "mail_state",         12),
    ("Mail ZIP",            "mail_zip",           12),
    ("Situs Address",       "situs_address",      28),
    ("Situs City",          "situs_city",         18),
    ("County",              "county",             18),
    ("State",               "state_name",         16),
    ("Legal Description",   "legal_description",  55),
    # Assessor-specific fields (populated by county_gis source)
    ("Account No",          "account_no",         18),
    ("Actual Value",        "actual_value",       16),
    ("Assessed Value",      "assessed_value",     16),
    ("Acreage",             "acreage",            14),
    ("Sec/Twp/Rng",        "section_township_range", 18),
    ("Tax Year",            "tax_year",           12),
    ("Assessor Portal",     "assessor_portal",    32),
    # LineString metadata (blank for Point files)
    ("Line Name",           "line_name",          28),
    ("Line Designator",     "line_designator",    20),
    ("Begin Station",       "begin_station",      16),
    ("End Station",         "end_station",        16),
    ("Operator",            "operator",           22),
    ("Product Type",        "product_type",       18),
    ("Op. Status",          "operational_status", 16),
    # Status always last
    ("Data Source",         "data_source",        24),
    ("Lookup Status",       "status",             20),
]

_BLUE_DARK  = "1F4E79"
_BLUE_LIGHT = "D6E4F0"
_GREEN      = "E2EFDA"
_ORANGE     = "FCE4D6"
_YELLOW     = "FFF2CC"


def _header_style():
    return Font(bold=True, color="FFFFFF", size=11), PatternFill("solid", fgColor=_BLUE_DARK)


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def write_excel(rows: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parcel Records"
    ws.sheet_view.showGridLines = False

    hdr_font, hdr_fill = _header_style()
    border = _thin_border()

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 2):
        status = row.get("status", "")
        alt = (row_idx % 2 == 0)

        if "ERROR" in str(status):
            row_fill = PatternFill("solid", fgColor=_ORANGE)
        elif "DUPLICATE" in str(status):
            row_fill = PatternFill("solid", fgColor=_YELLOW)
        elif alt:
            row_fill = PatternFill("solid", fgColor=_BLUE_LIGHT)
        else:
            row_fill = PatternFill("solid", fgColor="FFFFFF")

        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val = row.get(key, "")
            if key in ("lat", "lng") and isinstance(val, float):
                val = round(val, 6)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=(key == "legal_description"))
            cell.border = border

        ws.row_dimensions[row_idx].height = 18

    # ── Auto-filter ──────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "KMZ → Assessor Export Summary"
    ws2["A1"].font = Font(bold=True, size=14, color=_BLUE_DARK)
    ws2["A3"] = "Run date:"
    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A4"] = "Total samples:"
    ws2["B4"] = len(rows)
    ws2["A5"] = "Unique parcels queried:"
    ws2["B5"] = sum(1 for r in rows if not r.get("deduped") and r.get("parcel_id"))
    ws2["A6"] = "Records found:"
    ws2["B6"] = sum(1 for r in rows if r.get("parcel_id"))
    ws2["A7"] = "Duplicates (skipped):"
    ws2["B7"] = sum(1 for r in rows if r.get("deduped"))
    ws2["A8"] = "Errors:"
    ws2["B8"] = sum(1 for r in rows if "ERROR" in str(r.get("status", "")))
    ws2["A9"] = "Data source:"
    ws2["B9"] = "County Assessor GIS (free)"
    for cell in ["A3", "A4", "A5", "A6", "A7", "A8", "A9"]:
        ws2[cell].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Extract parcel owner records for KMZ points/lines from county assessor GIS data.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("kmz_file", help="Path to KMZ (or KML) file")
    parser.add_argument(
        "--output", "-o",
        default="parcel_results.xlsx",
        help="Output Excel file path (default: parcel_results.xlsx)",
    )
    parser.add_argument(
        "--sample-interval",
        type=float,
        default=0.0,
        dest="sample_interval",
        help=(
            "Sampling interval in decimal degrees for LineString geometry. "
            "0.0 (default) = every vertex. 0.001 = ~100 m spacing."
        ),
    )
    args = parser.parse_args()

    # ── Parse KMZ ────────────────────────────────────────────────────────────
    print(f"\nParsing: {args.kmz_file}")
    try:
        points = parse_kmz(args.kmz_file, sample_interval_deg=args.sample_interval)
    except Exception as e:
        print(f"Failed to parse KMZ: {e}")
        sys.exit(1)

    geom_counts = {}
    for p in points:
        g = p.get("geometry_type", "Unknown")
        geom_counts[g] = geom_counts.get(g, 0) + 1
    summary_parts = [f"{v} {k}(s)" for k, v in geom_counts.items()]
    print(f"   {len(points)} sample(s) extracted  ({', '.join(summary_parts)})")

    if not points:
        print("No samples to process. Exiting.")
        sys.exit(0)

    print("   SOURCE: Free County GIS (no API key required)\n")

    # ── Lookup each point (with deduplication) ────────────────────────────
    def _print_progress(i, total, label):
        print(f"[{i + 1:>3}/{total}] {label}")

    rows_raw = lookup_parcels_deduped(
        points,
        progress_callback=_print_progress,
    )
    rows = [_flatten_meta(r) for r in rows_raw]

    # ── Write Excel ───────────────────────────────────────────────────────────
    print(f"\nWriting Excel -> {args.output}")
    write_excel(rows, args.output)

    found        = sum(1 for r in rows if r.get("parcel_id"))
    unique_q     = sum(1 for r in rows if not r.get("deduped") and r.get("parcel_id"))
    deduped      = sum(1 for r in rows if r.get("deduped"))
    errors       = sum(1 for r in rows if "ERROR" in str(r.get("status", "")))

    print(f"\n{'=' * 55}")
    print(f"  Samples processed    : {len(rows)}")
    print(f"  Unique parcels queried: {unique_q}")
    print(f"  Duplicates skipped   : {deduped}")
    print(f"  Records found        : {found}")
    print(f"  Errors               : {errors}")
    print(f"  Output file          : {args.output}")
    print(f"  Data source          : County Assessor GIS (free)")
    print(f"{'=' * 55}\n")


if __name__ == "__main__":
    main()
